#-*- coding:utf-8 -*-
import datetime
from openpyxl import load_workbook
from openpyxl import Workbook

wk = load_workbook('courses.xlsx')
student_sheet = wk['students']
time_sheet = wk['time']

def combine():
    combine_sheet = wk.create_sheet(title="combine")
    combine_sheet.append(['创建时间','课程名称','学习人数','学习时间'])
    for stu in student_sheet.values:
        if stu[2]!= "学习人数":
            for time in time_sheet.values:
                if time[1]== stu[1]:
                    combine_sheet.append(list(stu)+ [time[2]])
    wk.save('courses.xlsx')
def split():
    combine_sheet = wk['combine']
    split_name = []
    for item  in combine_sheet.values:
        if item[0] != "创建时间":
            split_name.append(item[0].strftime('%Y'))
   
    for name in set(split_name):
        wk_temp = Workbook()
        wk_temp.remove(wk_temp.active)
        ws = wk_temp.create_sheet(title="name")
        for item_by_year in combine_sheet.values:
            if item_by_year[0] != "创建时间":
                if item_by_year[0].strftime('%Y') == name :
                    ws.append(item_by_year)

        wk_temp.save('{}.xlsx'.format(name))
if __name__ == "__main__":
    combine()
    split()
    
        


